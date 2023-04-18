# Created: 2/12/2023. By kiyunlee
# Edited: 4/9/2023

# Note this program is designed for images with pixel size 2048x2048.
# If not 2048x2048, change "factor" in the program such that (px height/factor) and (px width/factor) is a integer.

# Note this program is designed for tiff images.
# For jpg or png image, slight modifications of the code are needed.
# Use the following code with modifications.
# import cv2
# img = cv2.imread("geeksforgeeks.png", cv2.IMREAD_COLOR) # load image
# imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) # convert to gray scale


# You may need to install packages required. Use the following methods.
# pip install opencv-python
# pip install customtkinter



# =============================================================================
# Calcium dynamics tracer
# =============================================================================
import time; begin = time.time(); runtime = 0  # measure how long the program to excute
import numpy as np
import pandas as pd
from tkinter import filedialog
import customtkinter as ctk
import sys
import os.path
import tifffile as tif
import cv2
import openpyxl


# add graphic user interface
root = ctk.CTk(); root.title("Calcium dynamics tracer"); root.geometry('400x400')

i, j = 60, 50 # x, y coordinate
k = 15 # fontsize

def openDialog(): # open file dialog to select an image
    window = ctk.CTk()
    global file
    file = filedialog.askopenfilenames(parent=window, filetypes = [("Image file", ".tif")])
    b = ctk.CTkButton(root, text="Selected", width = 100, fg_color='red', font=('Arial',k), command=openDialog); b.place(x=i+150, y=j) 

    window.destroy()
        
def apply():
    global s; s = int(c2.get()) # user input: stage number
    global c; c = [c3a.get(),c3b.get(),c3c.get(),c3d.get(),c3e.get()]
    root.quit()
    

l1 = ctk.CTkLabel(root, text = 'Image file(s)', font=('Arial',k)); l1.place(x=i, y=j)
b1 = ctk.CTkButton(root, text="Open", width = 100, font=('Arial',k), command=openDialog); b1.place(x=i+150, y=j) 

l2 = ctk.CTkLabel(root, text = 'Stage', font=('Arial',k)); l2.place(x=i, y=j+50) 
c2 = ctk.CTkComboBox(root, width=100, values=[' 1',' 2',' 3',' 4',' 5',' 6'], dropdown_hover_color='red'); c2.place(x=i+150, y=j+50)

l3 = ctk.CTkLabel(root, text = 'Channel(s)', font=('Arial',k)); l3.place(x=i, y=j+100)
c3a = ctk.CTkCheckBox(root, text = 'DAPI', font=('Arial',k), fg_color='red', hover_color='red'); c3a.place(x=i+150, y=j+100)
c3b = ctk.CTkCheckBox(root, text = '340 nm', font=('Arial',k), fg_color='red', hover_color='red'); c3b.place(x=i+150, y=j+130)
c3c = ctk.CTkCheckBox(root, text = '380 nm', font=('Arial',k), fg_color='red', hover_color='red'); c3c.place(x=i+150, y=j+160)
c3d = ctk.CTkCheckBox(root, text = 'FITC', font=('Arial',k), fg_color='red', hover_color='red'); c3d.place(x=i+150, y=j+190)
c3e = ctk.CTkCheckBox(root, text = 'DsRed', font=('Arial',k), fg_color='red', hover_color='red'); c3e.place(x=i+150, y=j+220)

b4 = ctk.CTkButton(root, text="Apply", width=100, font=('Arial',k), command=apply); b4.place(x=i+150, y=j+270)

l5 = ctk.CTkLabel(root, text = 'By kiyunlee', text_color="red", font=('Arial',int(k*2/3))); l5.place(x=i+70, y=j+330) 

root.mainloop(); root.destroy()







# measure height and width of image
img = tif.imread(file[0])
h, w = np.shape(img)[0], np.shape(img)[1] # image size (px)

if (h!=2048) | (w!=2048):
    print('\nThe image size is not 2048x2048.\nPlease adjust the size factor and rerun the program.')
    sys.exit()
factor=4 # necessary to optimize window size. Multiply x, y coordinates by factor


# check correct files and channels are selected
if (len(file)==1) & (sum(c)!=1):
    print('\nFor one channel, one file should be chosen.\nPlease rerun the program.')
    sys.exit()
elif (len(file)==2) & (sum(c)!=2):
    print('\nFor two channels, two files should be chosen.\nPlease rerun the program.')
    sys.exit()
elif sum(c) >= 3:
    print('\nThe program is set up for either one or two channels.\nPlease rerun the program.')
    sys.exit()







# split file directory, name, extension
directory = os.path.split(file[0])[0]; directory = directory+'/'


temp = os.path.split(file[0])[1]; temp = temp.split('_')
ext = temp[-1].split('.')[-1]; temp = temp[:-2]
fileName = []; fileName.append(temp); del temp

name = []; temp = ''
for i in range(np.shape(fileName)[-1]):
    temp += fileName[0][i]+'_'
    
name.append(temp); del temp


if sum(c)==2: # when multichannel
    temp = os.path.split(file[1])[1]; temp = temp.split('_')
    temp = temp[:-2]; fileName.append(temp); del temp

    temp = ''
    for i in range(np.shape(fileName)[-1]):
        temp += fileName[1][i]+'_'
    
    name.append(temp); del temp
    

    



# measure the number of frame
def fullName(name,s,i): # define file name by stage number, frame number   
    fullName = name + 's%s_' % s + 't%i' % i + '.' + ext
    
    return fullName

i=1
while os.path.exists(directory+fullName(name[0],s,i)):
    i += 1
    
numFrame = i-1







# load image sequence
imgGray = np.zeros((numFrame,h,w))

for i in range(1,numFrame+1):
    if sum(c)==1: # when single channel
        img = tif.imread(directory+fullName(name[0],s,i)) # read individual image
        img = img/256; img = img.astype('uint8') # convert 16 to 8 bit grayscale
        
    elif sum(c)==2: # when multi-channels
        img1 = tif.imread(directory+fullName(name[0],s,i)) # read individual image
        img2 = tif.imread(directory+fullName(name[1],s,i)) 
        img = img1/img2
    
    imgGray[i-1,:,:] = img # expand 2D to 3D array (image to image sequence)
    







# autocontrast function
def autocontrast_func(img, cutoff=0):
    '''
        same output as PIL.ImageOps.autocontrast
    '''
    n_bins = 256
    def tune_channel(ch):
        n = ch.size
        cut = cutoff * n // 100
        if cut == 0:
            high, low = ch.max(), ch.min()
        else:
            hist = cv2.calcHist([ch], [0], None, [n_bins], [0, n_bins])
            low = np.argwhere(np.cumsum(hist) > cut)
            low = 0 if low.shape[0] == 0 else low[0]
            high = np.argwhere(np.cumsum(hist[::-1]) > cut)
            high = n_bins - 1 if high.shape[0] == 0 else n_bins - 1 - high[0]
        if high <= low:
            table = np.arange(n_bins)
        else:
            scale = (n_bins - 1) / (high - low)
            offset = -low * scale
            table = np.arange(n_bins) * scale + offset
            table[table < 0] = 0
            table[table > n_bins - 1] = n_bins - 1
        table = table.clip(0, 255).astype(np.uint8)
        return table[ch]
    channels = [tune_channel(ch) for ch in cv2.split(img)]
    out = cv2.merge(channels)
    return out






# select cells
print('\nPlease select cells.\nTo finish, press any key.') # user input

# define function to get x, y coordinates from user
pos = []
def click_event(event, x, y, flags, params):

    if event == cv2.EVENT_LBUTTONDOWN:
        
        print(f'({x},{y})')
        pos.append([x,y])
        
        # i = np.shape(pos)[0]
        # cv2.putText(img, f'{i}', (x,y), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2) # numbering on images
            
        cv2.circle(img, (x,y), 2, (0,0,255), -1) # indicators on images
        cv2.imshow('Selection',img)
    


img = tif.imread(directory+fullName(name[0],s,1)) # display the first image
img = img/256; img = img.astype('uint8') # convert 16 to 8 bit grayscale
img = autocontrast_func(img, cutoff=1) # autocontrast
img = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB) # gray to rgb

img = cv2.resize(img,(np.int(np.shape(img)[1]/factor), np.int(np.shape(img)[0]/factor))) # necessary to resize figure
cv2.imshow('Selection',img)
cv2.setMouseCallback('Selection', click_event)
cv2.waitKey(0) # stop recording once typing

posCell = pos; posCell = factor*np.array(posCell)







# select backgrounds
print('\nPlease select backgrounds.\nTo finish, press any key.') # user input

pos = []

cv2.setMouseCallback('Selection', click_event)
cv2.waitKey(0)

posBg = pos; posBg = factor*np.array(posBg)







# trace raw F, F0 before background subtraction
px = 6 # empirically determined
F = []; error = False

for i in range(np.shape(posCell)[0]):
    
    if (posCell[i,1]-px<0) | (posCell[i,1]+px>np.shape(imgGray)[1]-1) | (posCell[i,0]-px<0) | (posCell[i,0]+px>np.shape(imgGray)[2]-1):
        
        error = True
        print(f'Cell {i+1} is out of bounds.')
    
    else:
        temp = imgGray[:, posCell[i,1]-px:posCell[i,1]+px+1, posCell[i,0]-px:posCell[i,0]+px+1]
        temp = np.array(temp); temp = np.mean(np.mean(temp, axis=-1), axis=-1) # average values along y then x axis
        F.append(temp)

    
F = np.array(F); del temp






# define F0 when resting state (usually first 10 frames)
i, j = 0, 10
F0 = F[:,i:j]






# trace background intensity
F_bg = []

for i in range(np.shape(posBg)[0]):
    
    if (posBg[i,1]-px<0) | (posBg[i,1]+px>np.shape(imgGray)[1]-1) | (posBg[i,0]-px<0) | (posBg[i,0]+px>np.shape(imgGray)[2]-1):
        
        error = True
        print(f'Background {i+1} is out of bounds.')
        
    else:
        temp = imgGray[:, posBg[i,1]-px:posBg[i,1]+px+1, posBg[i,0]-px:posBg[i,0]+px+1]
        temp = np.array(temp); temp = np.mean(np.mean(temp, axis=-1), axis=-1) # average values along frames
        F_bg.append(temp)


F_bg = np.array(F_bg); F_bg = np.mean(F_bg, axis=0); # average along background selections
del temp


if error: # if selection is out of bounds, then exit the program
    print('\nPlease rerun the program.')
    import sys; sys.exit() # exit program







# calculate F, F0 after background subtraction
F = F-np.transpose(F_bg)*np.ones((np.shape(F)[0],np.shape(F)[1]))

i, j = 0, 10
F_bgRest = np.transpose(F_bg[i:j])*np.ones((np.shape(F0)[0], np.shape(F0)[1])) # background at resting state
F0 = F0-F_bgRest; F0 = np.mean(F0, axis=-1) # average along frame




# calculate dF/F0
trace = np.zeros((np.shape(F)[0],np.shape(F)[1]));
for i in range(numFrame):
    trace[:,i] = (F[:,i]-F0)/F0
    





# save data in excel
rows = [f'cell{i}' for i in range(1, np.shape(posCell)[0]+1)]
cols = [f't{i}' for i in range(1, numFrame+1)]
df  = pd.DataFrame(trace, index = rows, columns = cols) # create dataframe


name = 'results.xlsx'
if not os.path.exists(directory+name): # if excel file does not exist, create
    wb = openpyxl.Workbook(); wb.save(directory+name)
    
book = openpyxl.load_workbook(directory+name)
# book = load_workbook(directory+name)
writer = pd.ExcelWriter(directory+name, engine = 'openpyxl'); writer.book = book
df.to_excel(writer, sheet_name=f'stage{s}_raw') # add sheet 
writer.close()
    


end = time.time(); runtime += end-begin;
print(f'\nTotal runtime of the program: {runtime} s')











# =============================================================================
# Ca dynamics patterns
# =============================================================================
# This is optional. You can use this program to generate plots, or use your own.
# Plots will be generated from Calcium tracer tracer you just run
# To run this section, the previous section "Calcium dynamics tracer" has to run first.

import matplotlib.pyplot as plt


ctk.set_default_color_theme("green")
root = ctk.CTk(); root.title("Calcium activity pattern generator"); root.geometry('400x200')

i, j = 60, 50 # x, y coordinate
k = 15 # fontsize

def openDialog(): # open file dialog to select an image
    window = ctk.CTk()
    
    global meta
    meta = filedialog.askopenfilename(parent=window, filetypes = [("Excel file", ".xlsx .xls .cvs")])
    b = ctk.CTkButton(root, text="Selected", width = 100, fg_color='red', font=('Arial',k), command=openDialog); b.place(x=i+150, y=j) 

    window.destroy()
    
def apply():
    root.quit()
    

l1 = ctk.CTkLabel(root, text = 'Meta file', font=('Arial',k)); l1.place(x=i, y=j)
b1 = ctk.CTkButton(root, text="Open", width = 100, font=('Arial',k), command=openDialog); b1.place(x=i+150, y=j) 

b2 = ctk.CTkButton(root, text="Apply", width=100, font=('Arial',k), command=apply); b2.place(x=i+150, y=j+70)

l3 = ctk.CTkLabel(root, text = 'by kiyunlee', text_color="red", font=('Arial',int(k*2/3))); l3.place(x=i+70, y=j+130) 

root.mainloop(); root.destroy()



data = pd.read_excel (meta, sheet_name ='Injections', header = None)
data = data.to_numpy();
inj = data[1:,0]; comp = data[1:,4]



# subplots. cell score
# scale bar
t = 180/((inj[-1]-inj[0])/(len(inj)-1)) # sec per frame 
xScale, yScale = 180, 1 # sec, dF/F0
xMargin, yMargin = 5, .2


numFrame, numCell = np.shape(trace)[1], np.shape(trace)[0]
numSubplot = 10
l = 1 # line width
for i in range(numCell):
    
    if i % numSubplot == 0:
        if i+numSubplot < numCell:
            fig = plt.figure(f"Cell {i+1} to {i+numSubplot}", figsize=(7.2, 9.6)); j = 1
        else:
            fig = plt.figure(f"Cell {i+1} to {numCell}", figsize=(7.2, 9.6)); j = 1
            
            
    ax = fig.add_subplot(numSubplot+1,1,j); j += 1
    ax.plot(range(1,numFrame+1), trace[i,:], color='k', linewidth=l)
    ax.text(1+xMargin, np.max(trace)/2, f'{i+1}', color = "blue")
    
    ax.spines['top'].set_visible(False); ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False); ax.spines['right'].set_visible(False)
    plt.yticks([])
    plt.xticks(inj)
    ax.set_xticklabels(['']*len(inj))
    ax.tick_params(axis='x', colors='red')
    
    ax.set_xlim(1-xMargin,numFrame+xMargin)
    ax.set_ylim(np.min(trace)-yMargin,np.max(trace)+yMargin)
    
    plt.gca().invert_yaxis()
    
    if (j == 11) | (i == numCell-1): # scale bar
        ax = fig.add_subplot(numSubplot+1,1,j)
        ax.plot([1,1+xScale/t], [np.max(trace)-yScale, np.max(trace)-yScale], color='k', linewidth=1.5*l)   # x scale bar
        ax.plot([1,1], [np.max(trace)-yScale, np.max(trace)], color='k', linewidth=1.5*l) # y scale bar
    
        ax.axis('off')
        ax.set_xlim(1-xMargin,numFrame+xMargin)
        ax.set_ylim(np.min(trace)-yMargin,np.max(trace)+yMargin)
        
    fig.tight_layout()
    








# =============================================================================
# Ca dynamics pattern from Excel
# =============================================================================
# This is optional. You can use this program to generate plots, or use your own.
# Plots will be generated from excel file Calcium dynamic tracer created.
# To run this section, the previous section "Calcium dynamics tracer" has to run first.


import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import filedialog
import customtkinter as ctk
import openpyxl
from openpyxl import load_workbook

# add graphic user interface
ctk.set_default_color_theme("green")
root = ctk.CTk(); root.title("Calcium activity pattern generator"); root.geometry('400x400')

i, j = 60, 50 # x, y coordinate
k = 15 # fontsize

def openDialogResult(): # open file dialog for result file
    window = ctk.CTk()
    
    global result
    result = filedialog.askopenfilename(parent=window, filetypes = [("Excel file", ".xlsx .xls .cvs")])
    b = ctk.CTkButton(root, text="Selected", width = 100, fg_color='red', font=('Arial',k), command=openDialogResult); b.place(x=i+150, y=j) 

    window.destroy()

def openDialogMeta(): # open file dialog for meta file
    window = ctk.CTk()
    
    global meta
    meta = filedialog.askopenfilename(parent=window, filetypes = [("Excel file", ".xlsx .xls .cvs")])
    b = ctk.CTkButton(root, text="Selected", width = 100, fg_color='red', font=('Arial',k), command=openDialogMeta); b.place(x=i+150, y=j+100) 

    window.destroy()
    
def apply():
    global s; s = int(c2.get()) # user input: stage number
    root.quit()
    

l1 = ctk.CTkLabel(root, text = 'Result file', font=('Arial',k)); l1.place(x=i, y=j)
b1 = ctk.CTkButton(root, text="Open", width = 100, font=('Arial',k), command=openDialogResult); b1.place(x=i+150, y=j) 

l2 = ctk.CTkLabel(root, text = 'Stage', font=('Arial',k)); l2.place(x=i, y=j+50) 
c2 = ctk.CTkComboBox(root, width=100, values=[' 1',' 2',' 3',' 4',' 5',' 6'], dropdown_hover_color='red'); c2.place(x=i+150, y=j+50)

l3 = ctk.CTkLabel(root, text = 'Meta file', font=('Arial',k)); l3.place(x=i, y=j+100)
b3 = ctk.CTkButton(root, text="Open", width = 100, font=('Arial',k), command=openDialogMeta); b3.place(x=i+150, y=j+100) 

b4 = ctk.CTkButton(root, text="Apply", width=100, font=('Arial',k), command=apply); b4.place(x=i+150, y=j+270)

l5 = ctk.CTkLabel(root, text = 'By kiyunlee', text_color="red", font=('Arial',int(k*2/3))); l5.place(x=i+70, y=j+330) 


root.mainloop(); root.destroy()




# load result file
data = pd.read_excel (result, sheet_name =f'stage{s}_raw', header = None)
data = data.to_numpy() 

trace = data[1:,1:]
numFrame, numCell = np.shape(trace)[1], np.shape(trace)[0]



# load meta file
data = pd.read_excel (meta, sheet_name ='Injections', header = None)
data = data.to_numpy();

inj, comp = data[1:,0], data[1:,4]; # injection frame, compound name




# create analysis sheet in excel
from openpyxl import load_workbook

rows = [f'cell{i}' for i in range(1, numCell+1)]
cols = [comp[i] for i in range(len(comp))]
df  = pd.DataFrame('',index = rows, columns = cols)

book = load_workbook(result)
writer = pd.ExcelWriter(result, engine = 'openpyxl'); writer.book = book
df.to_excel(writer, sheet_name=f'stage{s}_analysis')  
writer.close()


# scale bar
t = 180/((inj[-1]-inj[0])/(len(inj)-1)) # sec per frame 
xScale, yScale = 180, 1 # sec, dF/F0
xMargin, yMargin = 5, .2




# create subplots for scoring cells
numSubplot = 10
l = 1 # line width
for i in range(numCell):
    
    if i % numSubplot == 0:
        if i+numSubplot < numCell:
            fig = plt.figure(f"Cell {i+1} to {i+numSubplot}", figsize=(7.2, 9.6)); j = 1
        else:
            fig = plt.figure(f"Cell {i+1} to {numCell}", figsize=(7.2, 9.6)); j = 1
            
            
    ax = fig.add_subplot(numSubplot+1,1,j); j +=1
    ax.plot(range(1,numFrame+1), trace[i,:], color='k', linewidth=l)
    ax.text(1+xMargin, np.max(trace)*1/3, f'{i+1}', color = "blue")
    
    ax.spines['top'].set_visible(False); ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False); ax.spines['right'].set_visible(False)
    plt.yticks([])
    plt.xticks(inj)
    ax.set_xticklabels(['']*len(inj))
    ax.tick_params(axis='x', colors='red')
    
    ax.set_xlim(1-xMargin,numFrame+xMargin)
    ax.set_ylim(np.min(trace)-yMargin,np.max(trace)+yMargin)
    
    plt.gca().invert_yaxis()
    
    if (j == 11) | (i == numCell-1): # scale bar
        ax = fig.add_subplot(numSubplot+1,1,j)
        ax.plot([1,1+xScale/t], [np.max(trace)-yScale, np.max(trace)-yScale], color='k', linewidth=1.5*l)   # x scale bar
        ax.plot([1,1], [np.max(trace)-yScale, np.max(trace)], color='k', linewidth=1.5*l) # y scale bar
    
        ax.axis('off')
        ax.set_xlim(1-xMargin,numFrame+xMargin)
        ax.set_ylim(np.min(trace)-yMargin,np.max(trace)+yMargin)
        
    fig.tight_layout()


directory = 'C:/Users/kiyunlee/Downloads/'
plt.savefig(directory + 'temp.png', transparent=True)

        
        


# created animated subplot for selected cells
for i in range(numFrame):
    
    # fig = plt.figure(figsize=(6.4, 4.8))
    fig = plt.figure(figsize=(8,8))
        
    for j in range(len(cell)):
    
        ax = fig.add_subplot(len(cell)+1,1,j+1)
        ax.plot(range(1,i+2), traceLocal[j,0:i+1], linewidth=l, label='cell')

        # ax.axis('off')
        ax.set_xlim(1-xMargin,numFrame+xMargin)
        ax.set_ylim(np.min(traceLocal)-yMargin,np.max(traceLocal)+yMargin)
        
        # adding injection points
        ax.spines['top'].set_visible(False); ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False); ax.spines['right'].set_visible(False)
        plt.yticks([])
        plt.xticks(inj)
        ax.set_xticklabels(['']*len(inj))
        ax.tick_params(axis='x', colors='red')
        
    # scale bar
    ax = fig.add_subplot(len(cell)+1,1,len(cell)+1)
    ax.plot([1,1+xScale/t], [np.max(traceLocal)-yScale, np.max(traceLocal)-yScale], color='k', linewidth=1.5*l)   # x scale bar
    ax.plot([1,1], [np.max(traceLocal)-yScale, np.max(traceLocal)], color='k', linewidth=1.5*l) # y scale bar

    ax.axis('off')
    ax.set_xlim(1-xMargin,numFrame+xMargin)
    ax.set_ylim(np.min(traceLocal)-yMargin,np.max(traceLocal)+yMargin)

    fig.tight_layout()
    
    name = "t%i.png" % i
    directory = 'C:/Users/kiyunlee/Downloads/temp images/' # change this according to your folder path
    plt.savefig(directory + name, transparent=True, dpi = 100)

    plt.close()






# =============================================================================
# recruitment curve generator
# =============================================================================
# This is optional. You can use this program to generate plots, or use your own.
# Plots will be generated from excel file containing cell scores.



# create dose response curve from excel file of cell scores
import scipy.optimize as opt


# add graphic user interface
ctk.set_default_color_theme("green")
root = ctk.CTk(); root.title("Calcium activity pattern generator"); root.geometry('400x400')

i, j = 60, 50 # x, y coordinate
k = 15 # fontsize

def openDialogResult(): # open file dialog for result file
    window = ctk.CTk()
    
    global result
    result = filedialog.askopenfilename(parent=window, filetypes = [("Excel file", ".xlsx .xls .cvs")])
    b = ctk.CTkButton(root, text="Selected", width = 100, fg_color='red', font=('Arial',k), command=openDialogResult); b.place(x=i+150, y=j) 

    window.destroy()

def openDialogMeta(): # open file dialog for meta file
    window = ctk.CTk()
    
    global meta
    meta = filedialog.askopenfilename(parent=window, filetypes = [("Excel file", ".xlsx .xls .cvs")])
    b = ctk.CTkButton(root, text="Selected", width = 100, fg_color='red', font=('Arial',k), command=openDialogMeta); b.place(x=i+150, y=j+100) 

    window.destroy()
    
def apply():
    global s; s = int(c2.get()) # user input: stage number
    root.quit()
    

l1 = ctk.CTkLabel(root, text = 'Result file', font=('Arial',k)); l1.place(x=i, y=j)
b1 = ctk.CTkButton(root, text="Open", width = 100, font=('Arial',k), command=openDialogResult); b1.place(x=i+150, y=j) 

l2 = ctk.CTkLabel(root, text = 'Stage', font=('Arial',k)); l2.place(x=i, y=j+50) 
c2 = ctk.CTkComboBox(root, width=100, values=[' 1',' 2',' 3',' 4',' 5',' 6'], dropdown_hover_color='red'); c2.place(x=i+150, y=j+50)

l3 = ctk.CTkLabel(root, text = 'Meta file', font=('Arial',k)); l3.place(x=i, y=j+100)
b3 = ctk.CTkButton(root, text="Open", width = 100, font=('Arial',k), command=openDialogMeta); b3.place(x=i+150, y=j+100) 

b4 = ctk.CTkButton(root, text="Apply", width=100, font=('Arial',k), command=apply); b4.place(x=i+150, y=j+270)

l5 = ctk.CTkLabel(root, text = 'By kiyunlee', text_color="red", font=('Arial',int(k*2/3))); l5.place(x=i+70, y=j+330) 


root.mainloop(); root.destroy()



data = pd.read_excel (result, sheet_name ='final analysis', header = None)
data = data.to_numpy() 

conct = data[16,4:11]
response = 100*data[-1,4:11]


def drCurve(x,b,c,d,e):
    # b: hill slope
    # c: min response
    # d: max response
    # e: EC50
    return (c+(d-c)/(1+np.exp(b*(np.log(x)-np.log(e)))))



fs, l = 16, 1 # font size, line width
fig = plt.figure(figsize=(6.4, 4.8))
ax = fig.add_subplot()

ax.plot(conct, response, color='b', marker='o', linestyle='none', mfc='none', label='cell')
fitCoeff, covMatrix = opt.curve_fit(drCurve, conct, response)

x = np.linspace(min(conct)*0.9,max(conct)*1.1,256)
ax.plot(x, drCurve(x,*[fitCoeff[i] for i in range(len(fitCoeff))]), color='b', linestyle='solid')

plt.xscale("log")



ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False);
ax.set_xticks([1,10,100,1000])
ax.set_xticklabels(['1','10','100','1000'], font='arial', fontsize=fs)
ax.set_xlabel(r'Concentration ($\mu$M)', font='arial', fontsize=fs+4)
ax.set_yticks([0,20,40,60,80,100])
ax.set_yticklabels(["0","20","40","60","80","100"], font='arial', fontsize=fs)
ax.set_ylabel('% Activated cells', font='arial', fontsize=fs+4)

plt.tight_layout()

directory = 'C:/Users/kiyunlee/Downloads/'
plt.savefig(directory + 'temp.png', transparent=True)


